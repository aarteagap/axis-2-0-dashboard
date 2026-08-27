#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETL Módulo de Sobrecostos — Arándano (Hortifrut Perú / Grupo Rocío / HFE Berries)
Procesa Registro_de_Gastos_T2324/T2425/T2526 (temporadas cerradas, liquidadas)
y GAOPEX_T2627_Conceptos (temporada actual, en revisión, datos parciales).

Reglas de negocio (definidas por el usuario, ver conversación):
- Hojas incluidas: MARÍTIMO, AÉREO, JORDIE, HFE_BERRIES, GRUPO ROCIO/GR* (solo la
  hoja transaccional real, ver abajo).
- Hojas excluidas: TERRESTRE (no genera sobrecostos), EXPO - SIN PO / EXPO - NO
  ARANDANO (gastos de cancelación), EXPO - LIQ / LIQ ARÁNDANOS (maquinaria,
  reexportaciones, otros), Gastos Peru, y todas las hojas de soporte
  (Key, CECOS-CUENTAS, Cruce*, Maestro conceptos Gestion, BD*, Hoja1,
  Gastos<Via> de trabajo). La hoja "GR" es una tabla de referencia estática
  (56 filas idénticas en los 3 archivos) y NO es transaccional: se excluye.
- Vía (Tipo de Expedición): para HORTIFRUT PERÚ la vía es el NOMBRE de la hoja
  (MARÍTIMO/AÉREO). Para JORDIE / GRUPO ROCIO / HFE_BERRIES la vía viene del
  campo "Tipo Expedición" de cada fila (MARITIMO/AEREO/SEA/AIR/TERRESTRE...).
  TERRESTRE se excluye siempre (no tiene sobrecostos).
- Exportador (Shipper) normalizado a: Hortifrut Perú S.A.C., Inversiones Jordie
  S.A., Tal S.A. (Talsa), HFE Berries Perú S.A.C.
- Monto USD: se usa "Monto USD" si la hoja la trae; si no, y Moneda=USD se usa
  Monto; si Moneda=PEN se convierte con el tipo de cambio promedio anual
  referencial de la temporada (ver FX_ASSUMPTIONS).
- Fecha de Emisión: si la hoja no la trae (caso HFE_BERRIES en T23-24), se usa
  ETD como proxy (queda anotado en meta.notes).
- Clasificación OPEX: se construye una tabla maestra Concepto -> Agrupador OPEX
  a partir de la hoja "Maestro conceptos Gestion" (T24-25/T25-26), unificando
  conflictos menores por criterio de mayoría/consistencia.
- Filtro global de conceptos: por defecto se EXCLUYEN (des-marcados, pero
  seleccionables) los conceptos de gestión/documentales que no son sobrecostos
  operativos: gestión documental, senasa, servicio integral, comisión de
  agenciamiento, certificado de origen, fee de seguridad, flete internacional/
  nacional, manejo de plataforma, courier, visto bueno, inspección senasa,
  movilidad, movilidad de alimentación, multa, notas de crédito.
- Filtro global de semana del año (ISO week de la Fecha de Emisión / proxy).
- Temporada actual (T26-27 / GAOPEX_T2627): el archivo NO trae Vía ni Fecha de
  Despacho todavía (columna "Mode" vacía, no existe "Dispatch Date"); según lo
  indicado, esos campos se completarán cuando se entregue la base de bookings
  para cruce. Por ahora solo se agregan Supplier, Concepto (OPEX Expense
  Description), Clasificación OPEX (OPEX Expense Group) y Monto USD (Total
  Expense USD), SIN ubicarlos en la línea de tiempo semanal. Se deja alerta.
"""
import json, math, re, unicodedata
from collections import defaultdict
from datetime import datetime
import openpyxl

BASE = "/root/.claude/uploads/2fb4a3aa-b69c-5e44-bf67-dc7a2b1ec063"
FILES = {
    "T23-24": f"{BASE}/0f08170c-Registro_de_Gastos_T2324.xlsx",
    "T24-25": f"{BASE}/7ae9e26e-Registro_de_Gastos_T2425.xlsx",
    "T25-26": f"{BASE}/d4d641a1-Registro_de_Gastos_T2526.xlsx",
}
CURRENT_FILE = f"{BASE}/7d255d3e-GAOPEX_T2627_Conceptos.xlsm"

SEASON_LABELS = {
    "T23-24": "2023-2024", "T24-25": "2024-2025",
    "T25-26": "2025-2026", "T26-27": "2026-2027",
}

FX_ASSUMPTIONS = {
    "T23-24": 3.75, "T24-25": 3.75, "T25-26": 3.70, "T26-27": 3.65,
}

SEASON_ANCHOR = {
    "T23-24": datetime(2023, 5, 1), "T24-25": datetime(2024, 5, 1),
    "T25-26": datetime(2025, 5, 1), "T26-27": datetime(2026, 5, 1),
}


def to_campaign_week(d, season):
    if d is None:
        return None
    if isinstance(d, str):
        try:
            d = datetime.fromisoformat(d)
        except Exception:
            return None
    anchor = SEASON_ANCHOR[season]
    delta_days = (d - anchor).days
    return math.floor(delta_days / 7) + 1

SHEETS_BY_YEAR = {
    "T23-24": {"MARÍTIMO": "Marítimo", "AÉREO": "Aéreo", "JORDIE": None,
               "HFE_BERRIES": None},
    "T24-25": {"MARÍTIMO": "Marítimo", "AÉREO": "Aéreo", "JORDIE": None,
               "HFE_BERRIES": None},
    "T25-26": {"MARÍTIMO": "Marítimo", "AÉREO": "Aéreo",
               "GRUPO ROCIO": None, "HFE_BERRIES": None},
}
# hojas donde la vía se toma del NOMBRE de hoja (no del campo Tipo Expedición)
SHEET_IS_VIA = {"MARÍTIMO", "AÉREO"}

SHEETS_EXCLUDED = {
    "T23-24": ["TERRESTRE", "EXPO - NO ARANDANO", "LIQ ARÁNDANOS", "GR",
               "Cruce FBL1N", "CECOS-CUENTAS", "Cruce HFE", "BD", "Key"],
    "T24-25": ["TERRESTRE", "EXPO - NO ARANDANO", "LIQ ARÁNDANOS", "GR",
               "CECOS-CUENTAS", "Maestro conceptos Gestion", "BD", "Key"],
    "T25-26": ["TERRESTRE", "EXPO - SIN PO", "EXPO - LIQ", "GR",
               "Gastos Peru", "GastosMaritimos", "GastosTerrestre",
               "GastosGrupoRocio", "GastosHFEB", "GastosAereos",
               "CECOS-CUENTAS", "Maestro conceptos Gestion",
               "BD_Despachos1", "Key", "Hoja1"],
}


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def norm(s):
    if s is None:
        return ""
    return strip_accents(str(s)).strip().upper()


def find_col(header, *aliases):
    normed = [norm(h) for h in header]
    for alias in aliases:
        a = norm(alias)
        for i, h in enumerate(normed):
            if h == a:
                return i
    for alias in aliases:
        a = norm(alias)
        for i, h in enumerate(normed):
            if a in h:
                return i
    return None


VIA_MAP = {
    "MARITIMO": "Marítimo", "SEA": "Marítimo",
    "AEREO": "Aéreo", "AIR": "Aéreo",
    "TERRESTRE": "Terrestre", "LAND": "Terrestre",
}


def map_via(v):
    return VIA_MAP.get(norm(v))


def map_shipper(v):
    n = norm(v)
    if "HORTIFRUT" in n or n == "HFPE":
        return "Hortifrut Perú S.A.C."
    if "JORDIE" in n:
        return "Inversiones Jordie S.A."
    if "TAL S.A" in n or "TALSA" in n or n == "TAL S.A.":
        return "Tal S.A. (Talsa)"
    if "HFE BERRIES" in n or "HFE_BERRIES" in n or "HFE  BERRIES" in n:
        return "HFE Berries Perú S.A.C."
    return "Sin especificar"


# ---------------------------------------------------------------------------
# Tabla maestra Concepto -> Agrupador OPEX (desde "Maestro conceptos Gestion")
# ---------------------------------------------------------------------------
def build_opex_master():
    mapping = {}
    for path in [FILES["T24-25"], FILES["T25-26"]]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Maestro conceptos Gestion"]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        pairs = [(i - 1, i) for i, h in enumerate(header) if h == "Agrupador"]
        for r in rows[1:]:
            for ci, gi in pairs:
                if ci < len(r) and gi < len(r) and r[ci] and r[gi]:
                    c = norm(r[ci])
                    g = str(r[gi]).strip()
                    mapping.setdefault(c, {}).setdefault(g, 0)
                    mapping[c][g] += 1
    # resolver conflictos por mayoría
    resolved = {}
    for c, groups in mapping.items():
        resolved[c] = sorted(groups.items(), key=lambda kv: -kv[1])[0][0]
    return resolved


OPEX_MASTER = build_opex_master()


def opex_for(concepto):
    c = norm(concepto)
    if c in OPEX_MASTER:
        return OPEX_MASTER[c]
    # heurística por palabra clave si el concepto no está en la maestra
    if "FLETE" in c and "NACION" in c:
        return "FLETE NACIONAL"
    if "FLETE" in c and "INTERNAC" in c:
        return "FLETE EXPORTACIÓN"
    if "SENASA" in c or "FITOSANIT" in c:
        return "GESTIÓN FITOSANITARIA"
    if "ADUANA" in c:
        return "GESTIÓN DE ADUANAS"
    if "SOBREESTADIA" in c or "ALMACEN" in c:
        return "ALMACEN TEMPORAL"
    if "INSULADO" in c or "MANTA" in c or "GEL PACK" in c or "REPALETIZ" in c or "MALLA" in c:
        return "ACONDICIONAMIENTO"
    if "PLATAFORMA" in c:
        return "PLATAFORMA DE GESTIÓN"
    return "OTROS"


# ---------------------------------------------------------------------------
# Conceptos NO-sobrecosto (des-marcados por defecto en el filtro global)
# ---------------------------------------------------------------------------
DEFAULT_EXCLUDE_PATTERNS = [
    "GESTION DOCUMENTAL", "SENASA", "SERVICIO INTEGRAL",
    "COMISION DE AG", "CERTIFICADO DE ORIGEN", "FEE DE SEGURIDAD",
    "FLETE INTERNACIONAL", "FLETE NACIONAL", "MANEJO DE PLATAFORMA",
    "COURIER", "VISTO BUENO", "MOVILIDAD", "MULTA", "NOTA DE CREDITO",
]


def is_default_active(concepto):
    c = norm(concepto)
    return not any(p in c for p in DEFAULT_EXCLUDE_PATTERNS)


def to_iso_week(d):
    if d is None:
        return None
    if isinstance(d, str):
        try:
            d = datetime.fromisoformat(d)
        except Exception:
            return None
    return d.isocalendar()[1]


def to_usd(monto, moneda, monto_usd, season):
    if monto_usd not in (None, "", 0):
        try:
            return float(monto_usd)
        except Exception:
            pass
    if monto in (None, ""):
        return 0.0
    try:
        monto = float(monto)
    except Exception:
        return 0.0
    m = norm(moneda)
    if m == "USD" or m == "":
        return monto
    if m == "PEN" or m == "S/" or m == "SOLES":
        return monto / FX_ASSUMPTIONS[season]
    return monto


# ---------------------------------------------------------------------------
# Extracción por hoja
# ---------------------------------------------------------------------------
def extract_sheet(ws, sheet_name, season):
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col = lambda *a: find_col(header, *a)

    i_exportador = col("Exportador")
    i_proveedor = col("Proveedor", "Supplier")
    i_fecha = col("Fecha Emision", "Fecha de Emision")
    i_fecha_fallback = col("ETD")
    i_concepto = col("Concepto")
    i_tipo_exp = col("Tipo Expedicion")
    i_booking = col("Booking")
    i_monto = col("Monto")
    i_moneda = col("Moneda")
    i_monto_usd = col("Monto USD")

    out = []
    n_no_fecha = 0
    for r in rows_iter:
        if i_concepto is None or r[i_concepto] is None:
            continue
        concepto = str(r[i_concepto]).strip()
        if not concepto:
            continue

        if sheet_name in SHEET_IS_VIA:
            via = "Marítimo" if sheet_name == "MARÍTIMO" else "Aéreo"
        else:
            via = map_via(r[i_tipo_exp]) if i_tipo_exp is not None else None
        if via != "Marítimo" and via != "Aéreo":
            continue  # excluye Terrestre / vía desconocida

        exportador_raw = r[i_exportador] if i_exportador is not None else sheet_name
        shipper = map_shipper(exportador_raw)

        fecha = r[i_fecha] if i_fecha is not None else None
        used_fallback = False
        if fecha is None and i_fecha_fallback is not None:
            fecha = r[i_fecha_fallback]
            used_fallback = True
        if fecha is None:
            n_no_fecha += 1
            continue
        week = to_iso_week(fecha)
        if week is None:
            n_no_fecha += 1
            continue
        cweek = to_campaign_week(fecha, season)

        booking = r[i_booking] if i_booking is not None else None
        booking = str(booking).strip() if booking not in (None, "", "(en blanco)") else "Sin Booking"

        proveedor = r[i_proveedor] if i_proveedor is not None else None

        monto = r[i_monto] if i_monto is not None else None
        moneda = r[i_moneda] if i_moneda is not None else None
        monto_usd_raw = r[i_monto_usd] if i_monto_usd is not None else None
        usd = to_usd(monto, moneda, monto_usd_raw, season)

        out.append({
            "season": season, "week": week, "cweek": cweek, "via": via, "shipper": shipper,
            "concepto": concepto, "opex": opex_for(concepto),
            "booking": booking, "proveedor": str(proveedor) if proveedor else "",
            "usd": round(usd, 2), "used_fecha_fallback": used_fallback,
        })
    return out, n_no_fecha


def load_season(season):
    path = FILES[season]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    notes = []
    total_no_fecha = 0
    for sheet_name in SHEETS_BY_YEAR[season]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        recs, n_no_fecha = extract_sheet(ws, sheet_name, season)
        records.extend(recs)
        total_no_fecha += n_no_fecha
        if any(r["used_fecha_fallback"] for r in recs):
            notes.append(f"{season}/{sheet_name}: sin 'Fecha Emisión' en la fuente; se usó ETD como proxy.")
    if total_no_fecha:
        notes.append(f"{season}: {total_no_fecha} filas descartadas por no tener fecha utilizable.")
    return records, notes


# ---------------------------------------------------------------------------
# Temporada actual (GAOPEX_T2627) — parcial, sin Vía/Fecha de Despacho aún
# ---------------------------------------------------------------------------
def load_current_season():
    wb = openpyxl.load_workbook(CURRENT_FILE, read_only=True, data_only=True)
    ws = wb["OPEX_EXPENSES_2627"]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col = lambda *a: find_col(header, *a)
    i_shipper = col("Shipper")
    i_supplier = col("Supplier")
    i_concepto = col("OPEX Expense Description")
    i_grupo = col("OPEX Expense Group")
    i_usd = col("Total Expense USD")
    i_booking = col("Booking")
    i_mode = col("Mode")

    by_concepto = defaultdict(lambda: {"usd": 0.0, "n": 0})
    by_grupo = defaultdict(lambda: {"usd": 0.0, "n": 0})
    by_supplier = defaultdict(lambda: {"usd": 0.0, "n": 0})
    by_shipper = defaultdict(lambda: {"usd": 0.0, "n": 0})
    total_usd, n = 0.0, 0
    n_with_mode = 0
    n_rows_total = 0
    total_usd_all_rows = 0.0

    for r in rows_iter:
        n_rows_total += 1
        usd_all = r[i_usd] if i_usd is not None else 0
        try:
            total_usd_all_rows += float(usd_all) if usd_all not in (None, "") else 0.0
        except Exception:
            pass
        concepto = r[i_concepto] if i_concepto is not None else None
        if concepto is None:
            continue
        usd = r[i_usd] if i_usd is not None else 0
        try:
            usd = float(usd) if usd not in (None, "") else 0.0
        except Exception:
            usd = 0.0
        grupo = str(r[i_grupo]).strip() if i_grupo is not None and r[i_grupo] else "SIN CLASIFICAR"
        supplier = str(r[i_supplier]).strip() if i_supplier is not None and r[i_supplier] else "Sin especificar"
        shipper = map_shipper(r[i_shipper]) if i_shipper is not None else "Sin especificar"
        if i_mode is not None and r[i_mode]:
            n_with_mode += 1

        by_concepto[str(concepto).strip()]["usd"] += usd
        by_concepto[str(concepto).strip()]["n"] += 1
        by_grupo[grupo]["usd"] += usd
        by_grupo[grupo]["n"] += 1
        by_supplier[supplier]["usd"] += usd
        by_supplier[supplier]["n"] += 1
        by_shipper[shipper]["usd"] += usd
        by_shipper[shipper]["n"] += 1
        total_usd += usd
        n += 1

    top_suppliers = sorted(by_supplier.items(), key=lambda kv: -kv[1]["usd"])[:20]

    return {
        "status": "en_revision",
        "total_usd": round(total_usd, 2),
        "n_lineas": n,
        "n_rows_total": n_rows_total,
        "n_rows_sin_clasificar": n_rows_total - n,
        "total_usd_all_rows": round(total_usd_all_rows, 2),
        "n_con_via_asignada": n_with_mode,
        "by_concepto": [{"concepto": k, "usd": round(v["usd"], 2), "n": v["n"]}
                         for k, v in sorted(by_concepto.items(), key=lambda kv: -kv[1]["usd"])],
        "by_opex": [{"opex": k, "usd": round(v["usd"], 2), "n": v["n"]}
                    for k, v in sorted(by_grupo.items(), key=lambda kv: -kv[1]["usd"])],
        "by_shipper": [{"shipper": k, "usd": round(v["usd"], 2), "n": v["n"]}
                       for k, v in sorted(by_shipper.items(), key=lambda kv: -kv[1]["usd"])],
        "top_proveedores": [{"proveedor": k, "usd": round(v["usd"], 2), "n": v["n"]}
                            for k, v in top_suppliers],
    }


def main():
    all_records = []
    all_notes = []
    for season in ["T23-24", "T24-25", "T25-26"]:
        recs, notes = load_season(season)
        all_records.extend(recs)
        all_notes.extend(notes)
        print(f"{season}: {len(recs)} líneas extraídas, total USD {sum(r['usd'] for r in recs):,.0f}")

    # -- agregación semanal (season, week, via, shipper, opex, concepto) --
    weekly = defaultdict(lambda: {"usd": 0.0, "n": 0})
    for r in all_records:
        key = (r["season"], r["week"], r["cweek"], r["via"], r["shipper"], r["opex"], r["concepto"])
        weekly[key]["usd"] += r["usd"]
        weekly[key]["n"] += 1

    weekly_rows = []
    for (s, w, cw, v, ex, o, c), agg in weekly.items():
        weekly_rows.append({
            "s": s, "w": w, "cw": cw, "v": v, "ex": ex, "o": o, "c": c,
            "usd": round(agg["usd"], 2), "n": agg["n"],
            "active": is_default_active(c),
        })

    # -- lista maestra de conceptos con default activo/inactivo --
    concept_totals = defaultdict(lambda: {"usd": 0.0, "n": 0, "opex": None})
    for r in all_records:
        concept_totals[r["concepto"]]["usd"] += r["usd"]
        concept_totals[r["concepto"]]["n"] += 1
        concept_totals[r["concepto"]]["opex"] = r["opex"]
    concept_list = [
        {"concepto": c, "opex": v["opex"], "usd": round(v["usd"], 2), "n": v["n"],
         "defaultActive": is_default_active(c)}
        for c, v in sorted(concept_totals.items(), key=lambda kv: -kv[1]["usd"])
    ]

    # -- agregación por booking (season, via, shipper, booking) con desglose OPEX --
    bookings = defaultdict(lambda: {"usd": 0.0, "n": 0, "opex": defaultdict(float),
                                     "proveedores": set()})
    for r in all_records:
        key = (r["season"], r["via"], r["shipper"], r["booking"])
        b = bookings[key]
        b["usd"] += r["usd"]
        b["n"] += 1
        b["opex"][r["opex"]] += r["usd"]
        if r["proveedor"]:
            b["proveedores"].add(r["proveedor"])

    booking_rows = []
    for (s, v, ex, bk), agg in bookings.items():
        booking_rows.append({
            "s": s, "v": v, "ex": ex, "bk": bk,
            "usd": round(agg["usd"], 2), "n": agg["n"],
            "opex": {k: round(val, 2) for k, val in agg["opex"].items()},
            "np": len(agg["proveedores"]),
        })
    booking_rows.sort(key=lambda b: -b["usd"])

    # -- proveedores top por temporada --
    prov_totals = defaultdict(lambda: {"usd": 0.0, "n": 0})
    for r in all_records:
        if not r["proveedor"]:
            continue
        key = (r["season"], r["proveedor"])
        prov_totals[key]["usd"] += r["usd"]
        prov_totals[key]["n"] += 1
    top_proveedores = defaultdict(list)
    for (s, p), agg in prov_totals.items():
        top_proveedores[s].append({"proveedor": p, "usd": round(agg["usd"], 2), "n": agg["n"]})
    for s in top_proveedores:
        top_proveedores[s] = sorted(top_proveedores[s], key=lambda x: -x["usd"])[:25]

    current = load_current_season()

    total_usd_by_season = defaultdict(float)
    n_by_season = defaultdict(int)
    n_bookings_by_season = defaultdict(set)
    for r in all_records:
        total_usd_by_season[r["season"]] += r["usd"]
        n_by_season[r["season"]] += 1
        n_bookings_by_season[r["season"]].add(r["booking"])

    data = {
        "meta": {
            "seasons": [
                {"key": s, "label": SEASON_LABELS[s], "closed": s != "T26-27"}
                for s in ["T23-24", "T24-25", "T25-26", "T26-27"]
            ],
            "vias": ["Marítimo", "Aéreo"],
            "exportadores": ["Hortifrut Perú S.A.C.", "Inversiones Jordie S.A.",
                              "Tal S.A. (Talsa)", "HFE Berries Perú S.A.C."],
            "opex_groups": sorted({c["opex"] for c in concept_list}),
            "fx_assumptions": {**FX_ASSUMPTIONS, "note": (
                "Tipo de cambio referencial (S/ por US$), promedio anual, usado "
                "únicamente cuando la fila no trae 'Monto USD' calculado en la "
                "fuente y la Moneda es PEN."
            )},
            "sheets_incluidas": ["MARÍTIMO", "AÉREO", "JORDIE", "GRUPO ROCIO",
                                   "HFE_BERRIES"],
            "sheets_excluidas": [
                "TERRESTRE (no genera sobrecostos)",
                "EXPO - SIN PO / EXPO - NO ARÁNDANO (gastos de cancelación, no sobrecostos)",
                "EXPO - LIQ / LIQ ARÁNDANOS (maquinaria, reexportaciones y otros)",
                "Gastos Perú (espejo consolidado de trabajo)",
                "GR (tabla de referencia estática, no transaccional)",
                "Hojas de soporte: Key, CECOS-CUENTAS, Cruce FBL1N, Cruce HFE, "
                "Maestro conceptos Gestión, BD/BD_Despachos1, Hoja1, "
                "Gastos<Vía> (copias de trabajo)",
            ],
            "default_excluded_concept_patterns": DEFAULT_EXCLUDE_PATTERNS,
            "n_lineas_total": sum(n_by_season.values()),
            "n_bookings_total": sum(len(v) for v in n_bookings_by_season.values()),
            "total_usd_total": round(sum(total_usd_by_season.values()), 2),
            "por_temporada": {
                s: {"usd": round(total_usd_by_season[s], 2), "n_lineas": n_by_season[s],
                    "n_bookings": len(n_bookings_by_season[s])}
                for s in ["T23-24", "T24-25", "T25-26"]
            },
            "notes": all_notes,
            "current_season_alert": (
                "GAOPEX_T2627 (temporada 2026-2027) está en proceso de revisión y "
                "aprobación por el equipo de liquidación. El archivo aún NO trae "
                "Vía de expedición (columna 'Mode' vacía) ni Fecha de Despacho "
                "('Dispatch Date'): esos campos se completarán cuando se entregue "
                "la base de bookings para hacer el cruce. Mientras tanto, esta "
                "temporada se muestra en un panel aparte con Proveedor, Concepto, "
                "Clasificación OPEX y Monto USD (Total Expense USD), SIN ubicarla "
                "en la línea de tiempo semanal ni en el filtro de Vía."
            ),
        },
        "weekly": weekly_rows,
        "concepts": concept_list,
        "bookings": booking_rows[:3000],
        "top_proveedores": top_proveedores,
        "current_season": current,
    }
    return data


if __name__ == "__main__":
    data = main()
    out_path = "/tmp/claude-0/-home-user-axis-2-0-dashboard/2fb4a3aa-b69c-5e44-bf67-dc7a2b1ec063/scratchpad/sobrecostos_data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print("OK ->", out_path)
    print("total_usd_total:", data["meta"]["total_usd_total"])
    print("por_temporada:", data["meta"]["por_temporada"])
    print("current_season total_usd:", data["current_season"]["total_usd"], "n_lineas:", data["current_season"]["n_lineas"])
